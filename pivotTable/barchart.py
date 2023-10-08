from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# wb = load_workbook(r"C:\Users\adity\Desktop\PythonScripting\PythonScripting\pivotTable\pivot_table.xlsx")
# sheet = wb['Report']

# min_col = wb.active.min_column
# max_col = wb.active.max_column
# min_row = wb.active.min_row
# max_row = wb.active.max_row

# # sheet['B8']='=SUM(B6:B7)'
# # sheet['B8'].style ='Currency'

# # for i in range(min_col+1,max_col+1):
# #     sheet[f'{get_column_letter(i)}{max_row+1}'] = f'=SUM({get_column_letter(i)}{min_row+1}:{get_column_letter(i)}{max_row})'
# #     sheet[f'{get_column_letter(i)}{max_row+1}'].style ='Currency'


# # Create a BarChart object
# barchart = BarChart()

# # Define data and categories references with integer values
# data = Reference(sheet, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row)
# categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row + 1, max_row=max_row)

# # Add data and categories to the chart
# barchart.add_data(data, titles_from_data=True)
# barchart.set_categories(categories)

# # Add the chart to the sheet at cell B12
# sheet.add_chart(barchart, "B12")

# # Customize the chart title and style
# barchart.title = "Sales Report"
# barchart.style = 5

# # Save the workbook
# wb.save(r"C:\Users\adity\Desktop\PythonScripting\PythonScripting\pivotTable\pivot_table.xlsx")


wb = load_workbook(r"C:\Users\adity\Desktop\PythonScripting\PythonScripting\pivotTable\Placement_Drive_2022_(B.Tech)1 - Copy.xlsx")

sheet = wb['Placement']

min_col = wb.active.min_column
max_col = 6
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(sheet, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=5)
categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row + 1, max_row=5)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# s = barchart.series[0]
# s.graphicalProperties.line.solidFill = "gg99000"
# s.graphicalProperties.solidFill = "ff9900" 

sheet.add_chart(barchart, "Z17")

barchart.title = "Placement Report"
barchart.style = 5

wb.save(r"C:\Users\adity\Desktop\PythonScripting\PythonScripting\pivotTable\Placement_Drive_2022_(B.Tech)1 - Copy.xlsx")
